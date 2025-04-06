from django.shortcuts import render, get_object_or_404, redirect
from .models import Employee
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from django.db.models import Count
from django.contrib.auth.decorators import user_passes_test
from .models import Employee
from .models import Employee

def is_superuser_or_staff(user):
    return user.is_superuser or user.is_staff
# Create your views here.
def index(request):
    employees = Employee.objects.all()
    return render(request, 'staff/index.html',{'employees': employees})

# @user_passes_test(is_superuser_or_staff)
# def employee_add(request):
#     if request.method == 'POST':
#         form = EmployeeForm(request.POST, request.FILES)
#         if form.is_valid():
#             form.save()
#             return redirect('staff:index')  # Redirect to the employee list
#     else:
#         form = EmployeeForm()
#     return render(request, 'staff/employee_form.html', {'form': form, 'title': 'Add Employee'})

# @user_passes_test(is_superuser_or_staff)
# def employee_edit(request, employee_id):
#     employee = get_object_or_404(Employee, pk=employee_id)
#     if request.method == 'POST':
#         form = EmployeeForm(request.POST, request.FILES, instance=employee)
#         if form.is_valid():
#             form.save()
#             return redirect('staff:index')  # Redirect to the employee list
#     else:
#         form = EmployeeForm(instance=employee)
#     return render(request, 'staff/employee_form.html', {'form': form, 'title': 'Edit Employee'})



@user_passes_test(is_superuser_or_staff)
def employee_detail(request, employee_id):
    employee = get_object_or_404(Employee, pk=employee_id)
    return render(request, 'staff/employee_detail.html', {'employee': employee})

def generate_report(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="employee_list.xlsx"'

    wb = Workbook()
    ws = wb.active

    # Create border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column headers and widths
    headers = [
        ('Personal No', 15), ('Name', 25), ('Father Name', 25),
        ('CNIC', 20), ('Designation', 20), ('Scale', 10),
        ('Post Status', 15), ('DOB', 15), ('Join Date', 15),
        ('Bank Info', 20), ('Phone', 15), ('Email', 25)
    ]

    # Write headers
    for col_num, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[chr(64+col_num)].width = width

    # Write data
    for row_num, employee in enumerate(Employee.objects.all(), 2):
        ws.cell(row=row_num, column=1, value=employee.personalno)
        ws.cell(row=row_num, column=2, value=employee.name)
        ws.cell(row=row_num, column=3, value=employee.fname)
        ws.cell(row=row_num, column=4, value=employee.cnic)
        ws.cell(row=row_num, column=5, value=employee.designation)
        ws.cell(row=row_num, column=6, value=employee.scale)
        ws.cell(row=row_num, column=7, value=employee.post_status)
        # Handle potential None for date_of_birth
        ws.cell(row=row_num, column=8, value=employee.date_of_birth.strftime('%Y-%m-%d') if employee.date_of_birth else '')
        # Handle potential None for date_of_first_joining (already handled)
        ws.cell(row=row_num, column=9, value=employee.date_of_first_joining.strftime('%Y-%m-%d') if employee.date_of_first_joining else '')
        ws.cell(row=row_num, column=10, value=f"{employee.accountno} ({employee.bank})")
        ws.cell(row=row_num, column=11, value=employee.phoneno)
        ws.cell(row=row_num, column=12, value=employee.email)

    # Apply formatting
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    wb.save(response)
    return response