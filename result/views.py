# views.py
from .models import Student, PrimaryResult, MiddleResult, HighResult,LEVEL_CHOICES,CLASS_CHOICES,RESULT_TYPE_CHOICES
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Q
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles import PatternFill
from django.db.models import Count
from django.db import models
from django.utils import timezone
import logging


logger = logging.getLogger(__name__)

# def home(request):
#     students=Student.objects.all()
#     context={
#         'students':students
#     }
#     return render(request, 'result/home.html',context)

def home(request):
    # Create mappings for display names
    level_names = dict(LEVEL_CHOICES)
    class_names = dict(CLASS_CHOICES)
    
    # Get summary data
    summary_data = (
        Student.objects
        .values('level', 'classs')
        .annotate(count=Count('id'))
        .order_by('level', 'classs')
    )
    
    # Organize data by level
    summary = {}
    for entry in summary_data:
        level_code = entry['level']
        class_code = entry['classs']
        
        if level_code not in summary:
            summary[level_code] = {
                'name': level_names.get(level_code, level_code),
                'classes': {},
                'total': 0
            }
            
        class_name = class_names.get(class_code, class_code)
        summary[level_code]['classes'][class_name] = entry['count']
        summary[level_code]['total'] += entry['count']
    
    # Convert to sorted list for template
    sorted_summary = []
    for level_code, level_name in LEVEL_CHOICES:
        if level_code in summary:
            sorted_classes = sorted(
                summary[level_code]['classes'].items(),
                key=lambda x: x[0]  # Sort classes by name
            )
            sorted_summary.append({
                'name': summary[level_code]['name'],
                'classes': sorted_classes,
                'total': summary[level_code]['total']
            })
    
    # Calculate grand total
    grand_total = sum(level['total'] for level in sorted_summary)
    
    context = {
        'summary': sorted_summary,
        'grand_total': grand_total,
        'grand_total': grand_total  # This must exist in the context
    }
    return render(request, 'result/home.html', context)

def student_list(request):
    selected_level = request.GET.get('level', '')
    selected_class = request.GET.get('classs', '')
    
    # Initialize empty queryset
    students = Student.objects.none()
    filters_applied = False

    # Check if any filters are provided (even if empty)
    if request.GET.get('level') or request.GET.get('classs'):
        filters_applied = True
        students = Student.objects.all()
        
        # Apply filters
        if selected_level:
            students = students.filter(level=selected_level)
        if selected_class:
            students = students.filter(classs=selected_class)

   # Excel Export Handler
    if 'export' in request.GET and filters_applied:
        try:
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            filename = f"students_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"
            
            # Custom Styles
            header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Column Headers
            headers = ['Name', "Father's Name", 'Roll No', 'Class', 'Level']
            ws.append(headers)
            
            # Apply Header Styles
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
            
            # Add Data Rows
            for student in students:
                ws.append([
                    student.name.strip(),
                    student.father_name.strip(),
                    student.roll_no.strip(),
                    student.get_classs_display(),
                    student.get_level_display(),
                ])
            
            # Set Column Widths
            column_widths = {
                'A': 30,  # Name
                'B': 30,  # Father's Name
                'C': 15,  # Roll No
                'D': 20,  # Class
                'E': 20   # Level
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Apply Data Alignment
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Freeze Header Row
            ws.freeze_panes = 'A2'
            
            # Save and return response
            wb.save(response)
            return response
        
        except Exception as e:
            logger.error(f"Excel export failed: {str(e)}")
            return HttpResponse("Error generating Excel file. Please try again later.", status=500)
    # Prepare context
    level_choices = Student._meta.get_field('level').choices
    class_choices = Student._meta.get_field('classs').choices

    return render(request, 'result/student_list.html', {
        'students': students if filters_applied else None,
        'LEVEL_CHOICES': level_choices,
        'CLASS_CHOICES': class_choices,
        'selected_level': selected_level,
        'selected_class': selected_class,
        'filters_applied': filters_applied,
    })

MODEL_MAP = {
    'primary': PrimaryResult,
    'middle': MiddleResult,
    'high': HighResult
}

def result_search(request):
    context = {
        'level_choices': [('primary', 'Primary'), ('middle', 'Middle'), ('high', 'High')],
        'class_choices': CLASS_CHOICES,
        'result_type_choices': RESULT_TYPE_CHOICES,
    }
    if request.method == 'POST':
        level = request.POST.get('level')
        classs = request.POST.get('classs').upper()
        roll_no = request.POST.get('roll_no')
        result_type = request.POST.get('result_type')  # Get the selected result type
        return redirect('result:result-detail', level=level, classs=classs, roll_no=roll_no, result_type=result_type)
    return render(request, 'result/search.html', context)

def result_detail(request, level, classs, roll_no,result_type):
    # result_type = request.GET.get('result_type') # Get result_type from the URL

    try:
        student = Student.objects.get(
            level=level,
            classs=classs,
            roll_no=roll_no
        )
    except Student.DoesNotExist:
        return render(request, 'result/student_not_found.html')

    results = []
    ResultModel = MODEL_MAP.get(student.level)
    if ResultModel and result_type:
        results = ResultModel.objects.filter(student=student, result_type=result_type).order_by('-year')
    elif ResultModel:
        results = ResultModel.objects.filter(student=student).order_by('-year', 'result_type')

    context = {
        'student': student,
        'results': results,
        'result_type_display': dict(RESULT_TYPE_CHOICES).get(result_type) if result_type else None,
    }
    return render(request, 'result/detail.html', context)

def class_results(request):
    # Get filter parameters
    level = request.GET.get('level')
    classs = request.GET.get('classs')
    result_type = request.GET.get('result_type')
    year = request.GET.get('year')

    students = Student.objects.none()
    results = []
    
    if all([level, classs, result_type, year]):
        students = Student.objects.filter(level=level, classs=classs).order_by('roll_no')
        
        # Get the appropriate result model
        ResultModel = MODEL_MAP.get(level)
        
        if ResultModel:
            # Prefetch results for all students
            results_qs = ResultModel.objects.filter(
                student__in=students,
                result_type=result_type,
                year=year
            ).select_related('student')
            
            # Create a mapping of student to result
            results_mapping = {result.student_id: result for result in results_qs}
            results = [results_mapping.get(student.id) for student in students]

    # Export handling
    export_format = request.GET.get('export')
    if export_format:
        # Prepare data for export
        data = []
        headers = ["Name", "Father's Name", "Roll No", "Class", "Total Marks", "Percentage", "Grade"]
        
        for student, result in zip(students, results):
            if result:
                row = [
                    student.name,
                    student.father_name,
                    student.roll_no,
                    student.get_classs_display(),
                    str(result.total_obtained),
                    f"{result.percentage}%",
                    result.grade
                ]
                data.append(row)
        
        if export_format == 'excel':
            return export_excel(headers, data)
        elif export_format == 'pdf':
            return export_pdf(headers, data)

    context = {
        'students': zip(students, results),
        'level': level,
        'classs': classs,
        'result_type': result_type,
        'year': year,
        'LEVEL_CHOICES': Student._meta.get_field('level').choices,
        'CLASS_CHOICES': Student._meta.get_field('classs').choices,
        'RESULT_TYPE_CHOICES': PrimaryResult._meta.get_field('result_type').choices,
    }
    return render(request, 'result/class_results.html', context)

def export_excel(headers, data):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="class_results.xlsx"'
    
    df = pd.DataFrame(data, columns=headers)
    df.to_excel(response, index=False)
    return response

def export_pdf(headers, data):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="class_results.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    
    # Create the table data
    table_data = [headers] + data
    
    # Create the table
    table = Table(table_data)
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    doc.build(elements)
    return response


def generate_report(request, report_type):
    # Configuration mapping for different report types
    REPORT_CONFIG = {
        'primary': {
            'model': PrimaryResult,
            'subjects': ['urdu', 'english', 'islamiat', 'gen_knowledge', 'math', 'science', 'nazra'],
            'title': 'Primary School Results',
            'filename': 'primary_results.xlsx'
        },
        'middle': {
            'model': MiddleResult,
            'subjects': ['urdu', 'english', 'islamiat', 'computer_or_Arabic', 'math', 'science', 'Al_Quran'],
            'title': 'Middle School Results',
            'filename': 'middle_results.xlsx'
        },
        'high': {
            'model': HighResult,
            'subjects': ['urdu', 'english', 'islamiat', 'computer', 'math', 
                        'biology', 'physics', 'pak_study', 'alquran'],
            'title': 'High School Results',
            'filename': 'high_results.xlsx'
        }
    }

    # Validate report type
    if report_type not in REPORT_CONFIG:
        return HttpResponse("Invalid report type", status=400)

    config = REPORT_CONFIG[report_type]
    results = config['model'].objects.select_related('student').all()

    # Create HTTP response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{config["filename"]}"'

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = config['title']

    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # Create columns
    columns = [
        ('Roll No', 15),
        ('Student Name', 25),
        ('Class', 15),
    ] + [(subject.upper(), 15) for subject in config['subjects']] + [
        ('Total Marks', 15),
        ('Obtained Marks', 15),
        ('Percentage', 15),
        ('Grade', 10)
    ]

    # Write headers
    for col_num, (header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Write data rows
    for row_num, result in enumerate(results, 2):
        # Student information
        ws.cell(row=row_num, column=1, value=result.student.roll_no)
        ws.cell(row=row_num, column=2, value=result.student.name)
        ws.cell(row=row_num, column=3, value=result.student.classs)

        # Subject marks
        subject_start_col = 4
        for idx, subject in enumerate(config['subjects']):
            col = subject_start_col + idx
            ws.cell(row=row_num, column=col, value=getattr(result, subject))

        # Result information
        total_col = subject_start_col + len(config['subjects'])
        ws.cell(row=row_num, column=total_col, value=result.total_marks)
        ws.cell(row=row_num, column=total_col+1, value=result.total_obtained)
        ws.cell(row=row_num, column=total_col+2, value=f"{result.percentage}%")
        ws.cell(row=row_num, column=total_col+3, value=result.grade)

        # Apply borders
        for col in range(1, len(columns)+1):
            ws.cell(row=row_num, column=col).border = thin_border

    wb.save(response)
    return response

def generate_class_report(request, class_name):
    # Get students in the specified class
    students = Student.objects.filter(classs=class_name).prefetch_related(
        'primaryresult_set', 
        'middleresult_set', 
        'highresult_set'
    )
    
    if not students.exists():
        return HttpResponse("No students found in this class", status=404)

    # Determine level from first student (assuming uniform class level)
    level = students.first().level
    SUBJECT_CONFIG = {
        'primary': PrimaryResult,
        'middle': MiddleResult,
        'high': HighResult
    }

    if level not in SUBJECT_CONFIG:
        return HttpResponse("Invalid class level", status=400)

    model = SUBJECT_CONFIG[level]
    results = model.objects.filter(student__in=students)

    # Create HTTP response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{class_name}_results.xlsx"'

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Class {class_name} Results"

    # Styling (same as generate_report)
    # ... [include the same styling code as above] ...

    # Get subject fields from model
    subjects = [field.name for field in model._meta.get_fields() 
               if isinstance(field, models.DecimalField) and field.name not in [
                   'total_marks', 'total_obtained', 'percentage']]

    # Create columns
    columns = [
        ('Roll No', 15),
        ('Student Name', 25),
    ] + [(subject.upper(), 15) for subject in subjects] + [
        ('Total Marks', 15),
        ('Obtained Marks', 15),
        ('Percentage', 15),
        ('Grade', 10)
    ]

    # Write headers
    # ... [same header writing code as generate_report] ...

    # Write data rows
    for row_num, result in enumerate(results, 2):
        # Student information
        ws.cell(row=row_num, column=1, value=result.student.roll_no)
        ws.cell(row=row_num, column=2, value=result.student.name)

        # Subject marks
        subject_start_col = 3
        for idx, subject in enumerate(subjects):
            col = subject_start_col + idx
            ws.cell(row=row_num, column=col, value=getattr(result, subject))

        # Result information
        # ... [same result info code as generate_report] ...

    wb.save(response)
    return response

   
    
# import csv
# from result.models import Student, CLASS_CHOICES

# # Define mappings for class to level
# PRIMARY_CLASSES = ['ECCE', '1', '2', '3', '4', '5']
# MIDDLE_CLASSES_PREFIXES = ['6IQBAL', '6QUAID', '7IQBAL', '7QUAID','8IQBAL','8QUAID','9IQBAL','9QUAID','10IQBAL','10QUAID']

# # Specify the CORRECTED path to your CSV file
# csv_file_path = r'E:\Employee\students\new without\students_data.csv'  # Using a raw string

# with open(csv_file_path, 'r', encoding='utf-8') as csvfile:
#     reader = csv.DictReader(csvfile)
#     students_to_create = []
#     for row in reader:
#         try:
#             classs = row['classs'].upper()  # Corrected header for class (two 's')
#             level = ''

#             if classs in PRIMARY_CLASSES:
#                 level = 'primary'
#             elif any(classs.startswith(prefix) for prefix in MIDDLE_CLASSES_PREFIXES):
#                 level = 'middle'
#             else:
#                 level = 'high'

#             student = Student(
#                 name=row['name'],
#                 father_name=row['father_name'],  # Corrected header for father's name
#                 roll_no=row['roll_no'],
#                 classs=classs,
#                 level=level,
#             )
#             students_to_create.append(student)

#         except KeyError as e:
#             print(f"Error processing row: {row} - Missing column: {e}")
#         except Exception as e:
#             print(f"Error processing row: {row} - {e}")

# if students_to_create:
#     Student.objects.bulk_create(students_to_create)
#     print(f"Successfully imported {len(students_to_create)} students.")
# else:
#     print("No new students to import.")